import csv
import io
from datetime import datetime, timedelta

import pytest
from openpyxl import load_workbook

from extensions import db
from models.tracked_device import (
    DeviceActivityLog,
    DeviceApplicationLog,
    TrackedDevice,
    TrackedDeviceAvailabilityEvent,
    TrackingSample,
)


pytestmark = pytest.mark.integration


def _create_report_device(mac_suffix: str) -> TrackedDevice:
    device = TrackedDevice(
        mac_address=f'AA:BB:CC:DD:EE:{mac_suffix}',
        device_name=f'Report-{mac_suffix}',
        employee_name=f'User {mac_suffix}',
        hostname=f'report-{mac_suffix.lower()}',
        ip_address='10.0.0.80',
        availability_status='online',
    )
    db.session.add(device)
    db.session.commit()
    return device


def _seed_tracking_window(device: TrackedDevice, start_utc: datetime, end_utc: datetime) -> None:
    first = start_utc + timedelta(minutes=3)
    second = end_utc - timedelta(minutes=4)
    for index, sample_time in enumerate((first, second), start=1):
        db.session.add(
            TrackingSample(
                device_id=device.id,
                idempotency_key=f'{device.id}:{index}:{sample_time.isoformat()}',
                received_at=sample_time,
                sampled_at=sample_time,
                integrity_status='verified' if index == 1 else 'partial',
            )
        )
        db.session.add(
            DeviceActivityLog(
                device_id=device.id,
                timestamp=sample_time,
                activity_type='keyboard' if index == 1 else 'mouse',
                event_count=5 + index,
            )
        )
        db.session.add(
            DeviceApplicationLog(
                device_id=device.id,
                timestamp=sample_time,
                application_name='Microsoft Word' if index == 1 else 'Google Chrome',
                duration=120 * index,
                status='active',
            )
        )
    db.session.add(
        TrackedDeviceAvailabilityEvent(
            device_id=device.id,
            observed_at=end_utc - timedelta(minutes=4),
            status='online',
            metrics_available=True,
        )
    )
    db.session.commit()


def test_workstation_reports_include_freshness_metadata(admin_client):
    end_utc = datetime.utcnow()
    start_utc = end_utc - timedelta(minutes=30)
    device = _create_report_device('D1')
    _seed_tracking_window(device, start_utc, end_utc)

    response = admin_client.get(
        f'/api/tracking/workstation/{device.id}/reports?from={start_utc.isoformat()}&to={end_utc.isoformat()}'
    )
    assert response.status_code == 200
    payload = response.get_json()

    freshness = payload['data']['freshness']
    assert freshness['source_basis'] == 'persisted_samples'
    assert freshness['sample_count'] == 2
    assert freshness['coverage_pct'] >= 10.0
    assert freshness['report_eligible'] is True
    assert 'last_sample_at' in freshness


def test_productivity_report_and_exports_include_freshness_fields(admin_client, app):
    app.config['ENABLE_PRODUCTIVITY_REPORT'] = True
    end_utc = datetime.utcnow()
    start_utc = end_utc - timedelta(minutes=30)
    device = _create_report_device('D2')
    _seed_tracking_window(device, start_utc, end_utc)

    params = f'?start={start_utc.isoformat()}&end={end_utc.isoformat()}&device_ids={device.id}'
    report_response = admin_client.get(f'/api/reports/productivity{params}')
    assert report_response.status_code == 200
    report_payload = report_response.get_json()

    assert report_payload['freshness_summary']['source_basis'] == 'persisted_samples'
    device_freshness = report_payload['freshness_summary']['devices'][str(device.id)]
    assert device_freshness['report_eligible'] is True
    assert device_freshness['sample_count'] == 2
    assert device_freshness['coverage_pct'] >= 10.0

    csv_response = admin_client.get(f'/api/reports/productivity/export{params}&format=csv')
    assert csv_response.status_code == 200
    csv_text = csv_response.data.decode('utf-8')
    reader = csv.DictReader(io.StringIO(csv_text))
    rows = list(reader)
    assert reader.fieldnames is not None
    assert 'Report Type' in reader.fieldnames
    assert 'Period Start' in reader.fieldnames
    assert 'Granularity' in reader.fieldnames
    assert 'Data Basis' in reader.fieldnames
    assert 'Freshness State' in reader.fieldnames
    assert 'Last Sample At' in reader.fieldnames
    assert 'Coverage %' in reader.fieldnames
    assert 'Sample Count' in reader.fieldnames
    assert 'Report Eligible' in reader.fieldnames
    assert rows
    assert rows[0]['Report Type'] == 'Productivity'
    assert rows[0]['Data Basis'] == 'persisted_samples'
    assert rows[0]['Report Eligible'] == 'Yes'

    xlsx_response = admin_client.get(f'/api/reports/productivity/export{params}&format=xlsx')
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.data))
    summary = workbook['Summary']
    summary_values = {(summary[f'A{row}'].value, summary[f'B{row}'].value) for row in range(1, summary.max_row + 1)}
    assert ('Data Basis', 'persisted_samples') in summary_values
    assert ('Fresh Devices', 1) in summary_values
    assert ('Report Type', 'Productivity') in summary_values
    assert 'Applications' in workbook.sheetnames
    assert 'Activity Summary' in workbook.sheetnames


def test_productivity_report_meta_omits_tracking_rollups_when_timescaledb_enabled(admin_client, app, monkeypatch):
    from services.timescaledb_service import TimescaleDBService

    app.config['ENABLE_PRODUCTIVITY_REPORT'] = True
    end_utc = datetime.utcnow()
    start_utc = end_utc - timedelta(days=2)
    device = _create_report_device('D3')
    _seed_tracking_window(device, start_utc, end_utc)
    monkeypatch.setattr(TimescaleDBService, 'is_timescaledb_enabled', lambda: True)

    response = admin_client.get(
        f'/api/reports/productivity?start={start_utc.isoformat()}&end={end_utc.isoformat()}&device_ids={device.id}'
    )
    assert response.status_code == 200
    payload = response.get_json()

    assert payload['meta']['source_tables'] == [
        'tracking_samples',
        'device_application_logs',
        'device_activity_logs',
    ]
    assert payload['meta']['freshness_sources'] == [
        'tracking_samples',
        'device_application_logs',
        'device_activity_logs',
    ]
    assert not any('tracking_hourly_rollups' in warning for warning in payload['meta']['completeness_warnings'])
    assert not any('tracking_daily_rollups' in warning for warning in payload['meta']['completeness_warnings'])
