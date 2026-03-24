"""
Server-side report export helpers.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from datetime import datetime, timedelta, timezone

logger = logging.getLogger(__name__)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(border_style="thin", color="4A5568"))
DATA_FONT = Font(name="Calibri", size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B2A4A")
ALT_ROW_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_INVALID_SHEET_CHARS_RE = re.compile(r"[:\\/?*\[\]]")
_CONTEXT_HEADERS = [
    "Report Type",
    "Period Start",
    "Period End",
    "Generated At",
    "Scope Type",
    "Scope ID",
    "Granularity",
    "Freshness State",
    "Data As Of",
]

_COLUMN_DISPLAY_NAMES = {
    "avg_uptime": "Avg Uptime %",
    "avg_latency": "Avg Latency (ms)",
    "avg_latency_ms": "Avg Latency (ms)",
    "avg_packet_loss": "Packet Loss %",
    "avg_cpu": "Avg CPU %",
    "max_cpu": "Max CPU %",
    "avg_mem": "Avg Memory %",
    "max_mem": "Max Memory %",
    "avg_disk": "Avg Disk %",
    "device_name": "Device",
    "device_ip": "IP Address",
    "device_type": "Type",
    "created_at": "Created At",
    "updated_at": "Updated At",
    "last_seen": "Last Seen",
    "is_acknowledged": "Acknowledged",
    "freshness_state": "Freshness",
    "coverage_pct": "Coverage %",
    "sample_count": "Samples",
    "report_eligible": "Report Eligible",
    "last_sample_at": "Last Sample At",
    "total_seconds": "Duration (sec)",
    "page_count_total": "Page Count",
}

_REPORT_TITLE_MAP = {
    "executive": "Executive Fleet Health",
    "operational": "Operational Activity",
    "device-health": "Device Health",
    "productivity": "Employee Productivity",
    "network": "Network Performance",
    "alerts": "Alert History",
    "maintenance-availability": "Maintenance & Availability",
    "security-compliance": "Security & Compliance",
    "inventory-assets": "Inventory & Assets",
    "tracking-operations": "Tracking Operations",
    "printer-operations": "Printer Operations",
}

_BRAND_FILL_DARK = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
_BRAND_FILL_MID = PatternFill(start_color="2D4A7A", end_color="2D4A7A", fill_type="solid")
_BRAND_FILL_LIGHT = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")


def _report_title(report_type: str) -> str:
    return _REPORT_TITLE_MAP.get(report_type, report_type.replace("-", " ").title())


def _brand_meta_text(report_data: dict, report_type: str):
    period = report_data.get("period") or {}
    meta = report_data.get("meta") or {}
    # Use plain ASCII separator to avoid encoding issues when opening in Excel
    title = f"Device Monitoring System  |  {_report_title(report_type)}"
    period_label = f"{period.get('start', '')} to {period.get('end', '')}"
    scope_id = meta.get("scope_id")
    scope_label = (
        f"{meta.get('scope_type', 'global')} : {scope_id}"
        if scope_id
        else meta.get("scope_type", "global")
    )
    freshness = meta.get("freshness_state", "") or "unknown"
    generated_at = meta.get("generated_at") or _utc_label()
    return title, period_label, scope_label, freshness, generated_at


def _write_brand_header_csv(text_buf: io.StringIO, report_data: dict, report_type: str) -> None:
    """Prepend APL TECHNO branded comment-header lines to the CSV string buffer."""
    _, period_label, scope_label, freshness, generated_at = _brand_meta_text(report_data, report_type)
    text_buf.write("# APL TECHNO \u2014 Network Monitoring System\n")
    text_buf.write(f"# {_report_title(report_type)}\n")
    text_buf.write(f"# Generated: {generated_at}  |  Period: {period_label}\n")
    text_buf.write(f"# Scope: {scope_label}  |  Freshness: {freshness}\n")
    text_buf.write("#\n")


def _write_brand_header_xlsx(ws, report_data: dict, report_type: str, num_cols: int) -> None:
    """Write APL TECHNO branding into worksheet rows 1-4.
    The caller must append a blank row 5 as a visual separator before writing data rows."""
    end_col = get_column_letter(max(num_cols, 2))
    title, period_label, scope_label, freshness, generated_at = _brand_meta_text(report_data, report_type)

    _META_FONT = Font(name="Calibri", size=10, color="333333")
    _META_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Row 1 — Company name (dark navy banner)
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value = "APL TECHNO"
    c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill = _BRAND_FILL_DARK
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Row 2 — Report title (medium blue)
    ws.merge_cells(f"A2:{end_col}2")
    c = ws["A2"]
    c.value = f"Device Monitoring System  |  {_report_title(report_type)}"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = _BRAND_FILL_MID
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    # Row 3 — Generated + Period (merged across all columns)
    ws.merge_cells(f"A3:{end_col}3")
    c = ws["A3"]
    c.value = f"Generated: {generated_at}  |  Period: {period_label}"
    c.font = _META_FONT
    c.fill = _META_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 18

    # Row 4 — Scope + Freshness (merged across all columns)
    ws.merge_cells(f"A4:{end_col}4")
    c = ws["A4"]
    c.value = f"Scope: {scope_label}  |  Freshness: {freshness}"
    c.font = _META_FONT
    c.fill = _META_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 18


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _utc_label() -> str:
    return datetime.now().strftime("%dd-%mm-%YY %H:%M")


_IST_TZ = timezone(timedelta(hours=5, minutes=30))


def _fmt_ist(val) -> str:
    """Format any timestamp/ISO string as IST for PDF display. Never outputs raw ISO 8601."""
    if val is None or val == "":
        return "—"
    if isinstance(val, str):
        if not val or val == "—":
            return "—"
        try:
            val = datetime.fromisoformat(val.replace("Z", "+00:00"))
        except (ValueError, TypeError):
            try:
                val = datetime.fromisoformat(val[:19])
            except Exception:
                return val[:16].replace("T", " ")
    if isinstance(val, datetime):
        if val.tzinfo is None:
            val = val.replace(tzinfo=timezone.utc)
        ist = val.astimezone(_IST_TZ)
        return ist.strftime("%d %b %Y %H:%M IST")
    return str(val)


def _sanitize_export_value(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        else:
            value = value.astimezone(timezone.utc)
        return value.isoformat().replace("+00:00", "Z")

    text = _ILLEGAL_XLSX_CHARS_RE.sub("", str(value))
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _parse_excel_datetime(value):
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text or len(text) < 10:
        return None
    candidate = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(candidate)
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        return parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def _excel_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float, datetime)):
        return _parse_excel_datetime(value) or value
    parsed = _parse_excel_datetime(value)
    if parsed is not None:
        return parsed
    return _sanitize_export_value(value)


def _report_context(report_data, report_type):
    period = report_data.get("period") or {}
    meta = report_data.get("meta") or {}
    return {
        "Report Type": report_type.replace("-", " ").title(),
        "Period Start": period.get("start", ""),
        "Period End": period.get("end", ""),
        "Generated At": meta.get("generated_at") or _utc_label(),
        "Scope Type": meta.get("scope_type", ""),
        "Scope ID": meta.get("scope_id", ""),
        "Granularity": meta.get("granularity", ""),
        "Freshness State": meta.get("freshness_state", ""),
        "Data As Of": meta.get("data_as_of", ""),
    }


def _apply_export_context(report_data, report_type, headers, rows):
    context = _report_context(report_data, report_type)
    contextual_headers = list(_CONTEXT_HEADERS)
    for header in headers:
        if header not in contextual_headers:
            contextual_headers.append(header)
    contextual_rows = [{**context, **row} for row in rows]
    return contextual_headers, contextual_rows


def _section_sheet_name(section, existing_names):
    base = str(section or "Data").replace("_", " ").strip().title() or "Data"
    base = _INVALID_SHEET_CHARS_RE.sub("", base)[:31] or "Data"
    if base not in existing_names:
        return base

    suffix = 2
    while True:
        candidate = f"{base[:28]}_{suffix}"
        if candidate not in existing_names:
            return candidate
        suffix += 1


def _group_rows_by_section(rows):
    grouped = {}
    ordered_sections = []
    for row in rows:
        section = row.get("Section") or "data"
        if section not in grouped:
            grouped[section] = []
            ordered_sections.append(section)
        grouped[section].append(row)
    return ordered_sections, grouped


def _sheet_title(ws, title, width=4):
    end_col = get_column_letter(max(2, width))
    ws.merge_cells(f"A1:{end_col}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_excel_number_format(cell, header):
    value = cell.value
    normalized = str(header or "").strip().lower()
    if isinstance(value, datetime):
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
        return
    if isinstance(value, int):
        cell.number_format = "#,##0"
        return
    if isinstance(value, float):
        if any(token in normalized for token in ("%", "pct", "uptime", "util", "coverage")):
            cell.number_format = "0.00"
        elif any(token in normalized for token in ("latency", "loss", "score", "confidence")):
            cell.number_format = "0.00"
        elif "bps" in normalized:
            cell.number_format = "#,##0.00"
        else:
            cell.number_format = "#,##0.00"


def _style_sheet_headers(ws, headers, header_row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER


def _autosize_worksheet(ws, headers, row_count, start_row=1):
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header or ""))
        for row_idx in range(start_row + 1, min(start_row + row_count + 1, start_row + 201)):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[ws.cell(row=start_row, column=col_idx).column_letter].width = min(max_len + 3, 40)


def _write_table_sheet(ws, headers, rows, start_row=1):
    _style_sheet_headers(ws, headers, header_row=start_row)

    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, header in enumerate(headers, 1):
            value = _excel_cell_value(row_data.get(header, ""))
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )
            cell.font = DATA_FONT
            _apply_excel_number_format(cell, header)
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    ws.freeze_panes = f"A{start_row + 1}"
    last_col = get_column_letter(len(headers))
    last_row = max(start_row + len(rows), start_row)
    ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"
    _autosize_worksheet(ws, headers, len(rows), start_row=start_row)


def _write_kpi_sheet(ws, title, items):
    _sheet_title(ws, title, width=4)
    start_row = 3
    for row_idx, (label, value) in enumerate(items, start=start_row):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = Alignment(vertical="top")
        value_cell = ws.cell(row=row_idx, column=2, value=_excel_cell_value(value))
        value_cell.font = DATA_FONT
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        _apply_excel_number_format(value_cell, label)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42


def _summary_pairs(report_data, report_type, row_count):
    period = report_data.get("period") or {}
    meta = report_data.get("meta") or {}
    pairs = [
        ("Report Type", report_type.replace("-", " ").title()),
        ("Period Start", period.get("start", "")),
        ("Period End", period.get("end", "")),
        ("Generated", meta.get("generated_at") or _utc_label()),
        ("Rows", row_count),
        ("Scope Type", meta.get("scope_type", "")),
        ("Scope ID", meta.get("scope_id", "")),
        ("Granularity", meta.get("granularity", "")),
        ("Freshness State", meta.get("freshness_state", "")),
        ("Data As Of", meta.get("data_as_of", "")),
        ("Max Source Lag (s)", meta.get("max_source_lag_seconds", "")),
        ("Source Tables", ", ".join(meta.get("source_tables") or [])),
        ("Freshness Sources", ", ".join(meta.get("freshness_sources") or [])),
    ]
    return pairs


def _write_titled_table_sheet(ws, title, headers, rows):
    _sheet_title(ws, title, width=len(headers))
    _write_table_sheet(ws, headers, rows, start_row=3)


def _build_executive_excel_sheets(wb, report_data):
    ws_overview = wb.create_sheet(title="Overview")
    sla_metrics = report_data.get("sla_metrics") or {}
    _write_kpi_sheet(
        ws_overview,
        "Executive Overview",
        [
            ("Uptime Score (%)", report_data.get("uptime_score")),
            ("Avg Latency (ms)", report_data.get("avg_latency")),
            ("Availability Basis", report_data.get("availability_basis")),
            ("Total Devices", report_data.get("total_devices")),
            ("MTTA (s)", sla_metrics.get("mtta_seconds")),
            ("MTTA", sla_metrics.get("mtta_human")),
        ],
    )

    ws_health = wb.create_sheet(title="Health Mix")
    health_rows = [
        {"Status": label, "Count": count}
        for label, count in (report_data.get("health_distribution") or {}).items()
    ]
    _write_titled_table_sheet(ws_health, "Health Distribution", ["Status", "Count"], health_rows)

    ws_devices = wb.create_sheet(title="Problem Devices")
    device_rows = []
    for item in (report_data.get("top_problematic") or []):
        uptime = item.get("uptime") or 0
        availability = "CRITICAL" if uptime < 95 else ("WARNING" if uptime < 99 else "OK")
        device_rows.append({
            "Device": item.get("name"),
            "IP": item.get("ip"),
            "Type": item.get("type"),
            "Uptime %": uptime,
            "Availability": availability,
        })
    _write_titled_table_sheet(
        ws_devices,
        "Problem Devices",
        ["Device", "IP", "Type", "Uptime %", "Availability"],
        device_rows,
    )
    # Colour-code the Availability column (col 5): CRITICAL=red, WARNING=amber, OK=green
    _AVAIL_FONT_COLORS = {"CRITICAL": "FF3B5C", "WARNING": "FFAA00", "OK": "00AA66"}
    avail_col_idx = 5
    for r in range(4, 4 + len(device_rows)):
        cell = ws_devices.cell(row=r, column=avail_col_idx)
        color = _AVAIL_FONT_COLORS.get(str(cell.value or "").upper())
        if color:
            cell.font = Font(name="Calibri", size=10, bold=True, color=color)


def _build_network_excel_sheets(wb, report_data):
    ws_overview = wb.create_sheet(title="Overview")
    mttr = report_data.get("mttr") or {}
    interface_count = len(report_data.get("bandwidth") or {})
    _write_kpi_sheet(
        ws_overview,
        "Network Overview",
        [
            ("Uptime Basis", report_data.get("uptime_basis")),
            ("Tracked Interfaces", interface_count),
            ("Uptime Rows", len(report_data.get("uptime_summary") or [])),
            ("MTTR (s)", mttr.get("seconds")),
            ("MTTR", mttr.get("human")),
            ("Resolved Incidents", mttr.get("total_incidents")),
        ],
    )

    ws_uptime = wb.create_sheet(title="Uptime Summary")
    uptime_rows = [
        {
            "Device": item.get("device_name"),
            "Avg Uptime %": item.get("avg_uptime"),
            "Avg Latency (ms)": item.get("avg_latency_ms"),
            "Avg Packet Loss %": item.get("avg_packet_loss"),
        }
        for item in (report_data.get("uptime_summary") or [])
    ]
    _write_titled_table_sheet(
        ws_uptime,
        "Uptime Summary",
        ["Device", "Avg Uptime %", "Avg Latency (ms)", "Avg Packet Loss %"],
        uptime_rows,
    )

    ws_bandwidth = wb.create_sheet(title="Bandwidth")
    bandwidth_rows = []
    for info in (report_data.get("bandwidth") or {}).values():
        for point in info.get("points") or []:
            bandwidth_rows.append(
                {
                    "Device": info.get("device_name"),
                    "Interface": info.get("interface"),
                    "Timestamp": point.get("ts"),
                    "RX (bps)": point.get("rx_bps"),
                    "TX (bps)": point.get("tx_bps"),
                    "RX Util %": point.get("rx_util"),
                    "TX Util %": point.get("tx_util"),
                }
            )
    _write_titled_table_sheet(
        ws_bandwidth,
        "Bandwidth Timeline",
        ["Device", "Interface", "Timestamp", "RX (bps)", "TX (bps)", "RX Util %", "TX Util %"],
        bandwidth_rows,
    )


def _build_alerts_excel_sheets(wb, report_data):
    ws_overview = wb.create_sheet(title="Overview")
    alerts = report_data.get("alerts") or []
    _write_kpi_sheet(
        ws_overview,
        "Alerts Overview",
        [
            ("Total Alerts", len(alerts)),
            ("Acknowledged Alerts", sum(1 for item in alerts if item.get("is_acknowledged"))),
            ("Resolved Alerts", sum(1 for item in alerts if item.get("resolved"))),
            ("TTA (s)", (report_data.get("tta") or {}).get("seconds")),
            ("TTA", (report_data.get("tta") or {}).get("human")),
            ("TTR (s)", (report_data.get("ttr") or {}).get("seconds")),
            ("TTR", (report_data.get("ttr") or {}).get("human")),
        ],
    )

    ws_alerts = wb.create_sheet(title="Alerts")
    alert_rows = [
        {
            "Timestamp": item.get("timestamp"),
            "Device": item.get("device_name"),
            "IP": item.get("device_ip"),
            "Severity": item.get("severity"),
            "Type": item.get("event_type"),
            "Message": item.get("message"),
            "Acknowledged": "Yes" if item.get("is_acknowledged") else "No",
            "Resolved": "Yes" if item.get("resolved") else "No",
            "Resolved At": item.get("resolved_at"),
        }
        for item in alerts
    ]
    _write_titled_table_sheet(
        ws_alerts,
        "Alert Detail",
        ["Timestamp", "Device", "IP", "Severity", "Type", "Message", "Acknowledged", "Resolved", "Resolved At"],
        alert_rows,
    )

    ws_severity = wb.create_sheet(title="Severity Mix")
    severity_rows = [
        {"Severity": severity, "Count": count}
        for severity, count in (report_data.get("severity_breakdown") or {}).items()
    ]
    _write_titled_table_sheet(ws_severity, "Severity Breakdown", ["Severity", "Count"], severity_rows)

    ws_trend = wb.create_sheet(title="Daily Trend")
    trend_rows = []
    for day, breakdown in (report_data.get("daily_trend") or {}).items():
        for severity, count in (breakdown or {}).items():
            trend_rows.append({"Date": day, "Severity": severity, "Count": count})
    _write_titled_table_sheet(ws_trend, "Daily Trend", ["Date", "Severity", "Count"], trend_rows)

    ws_top = wb.create_sheet(title="Top Devices")
    top_rows = [
        {
            "Device": item.get("device_name"),
            "IP": item.get("device_ip"),
            "Alert Count": item.get("alert_count"),
        }
        for item in (report_data.get("top_alerted_devices") or [])
    ]
    _write_titled_table_sheet(ws_top, "Top Alerted Devices", ["Device", "IP", "Alert Count"], top_rows)


def _flatten_report_rows(report_data, report_type):
    flatteners = {
        "device-health": _flatten_device_health,
        "productivity": _flatten_productivity,
        "network": _flatten_network,
        "alerts": _flatten_alerts,
        "executive": _flatten_executive,
        "operational": _flatten_operational,
        "maintenance-availability": _flatten_maintenance_availability,
        "security-compliance": _flatten_security_compliance,
        "inventory-assets": _flatten_inventory_assets,
        "tracking-operations": _flatten_tracking_operations,
        "printer-operations": _flatten_printer_operations,
    }
    flattener = flatteners.get(report_type)
    if flattener is None:
        return ["Section", "Key", "Value"], []
    return flattener(report_data or {})


def _flatten_executive(data):
    headers = ["Section", "Label", "Value", "Device", "IP", "Type", "Uptime %"]
    rows = [
        {"Section": "summary", "Label": "Uptime Score (%)", "Value": data.get("uptime_score")},
        {"Section": "summary", "Label": "Avg Latency (ms)", "Value": data.get("avg_latency")},
        {"Section": "summary", "Label": "MTTA", "Value": (data.get("sla_metrics") or {}).get("mtta_human")},
        {"Section": "summary", "Label": "Total Devices", "Value": data.get("total_devices")},
    ]
    for label, count in (data.get("health_distribution") or {}).items():
        rows.append({"Section": "health_distribution", "Label": label, "Value": count})
    for device in data.get("top_problematic") or []:
        rows.append(
            {
                "Section": "top_problematic",
                "Device": device.get("name"),
                "IP": device.get("ip"),
                "Type": device.get("type"),
                "Uptime %": device.get("uptime"),
            }
        )
    return headers, rows


def _flatten_operational(data):
    headers = ["Section", "Day", "Hour", "Count", "Type", "Detail", "Timestamp", "Device", "IP"]
    rows = []
    for day, hour, count in data.get("heatmap") or []:
        rows.append({"Section": "heatmap", "Day": day, "Hour": hour, "Count": count})
    for event in data.get("audit_log") or []:
        rows.append(
            {
                "Section": "audit_log",
                "Type": event.get("event_type"),
                "Detail": event.get("message"),
                "Timestamp": event.get("timestamp"),
            }
        )
    for device in data.get("new_devices") or []:
        rows.append(
            {
                "Section": "new_devices",
                "Device": device.get("device_name"),
                "IP": device.get("device_ip"),
                "Type": device.get("device_type"),
                "Timestamp": device.get("created_at"),
            }
        )
    return headers, rows


def _flatten_device_health(data):
    headers = [
        "Section",
        "Device",
        "Timestamp",
        "CPU %",
        "Memory %",
        "Disk %",
        "Net In (bps)",
        "Net Out (bps)",
        "Avg CPU %",
        "Max CPU %",
        "Avg Memory %",
        "Max Memory %",
        "Avg Disk %",
        "Samples",
    ]
    rows = []
    for item in data.get("summary") or []:
        rows.append(
            {
                "Section": "summary",
                "Device": item.get("device_name"),
                "Avg CPU %": item.get("avg_cpu"),
                "Max CPU %": item.get("max_cpu"),
                "Avg Memory %": item.get("avg_mem"),
                "Max Memory %": item.get("max_mem"),
                "Avg Disk %": item.get("avg_disk"),
                "Samples": item.get("samples"),
            }
        )
    for _, info in (data.get("time_series") or {}).items():
        for point in info.get("points") or []:
            rows.append(
                {
                    "Section": "time_series",
                    "Device": info.get("device_name"),
                    "Timestamp": point.get("ts"),
                    "CPU %": point.get("cpu"),
                    "Memory %": point.get("mem"),
                    "Disk %": point.get("disk"),
                    "Net In (bps)": point.get("net_in"),
                    "Net Out (bps)": point.get("net_out"),
                }
            )
    return headers, rows


def _flatten_productivity(data):
    headers = [
        "Section",
        "Device",
        "Employee",
        "Application",
        "Category",
        "Duration (sec)",
        "Sessions",
        "Data Basis",
        "Metric",
        "Value",
        "Freshness State",
        "Last Sample At",
        "Coverage %",
        "Sample Count",
        "Report Eligible",
    ]
    rows = []
    freshness_summary = data.get("freshness_summary") or {}
    freshness_devices = freshness_summary.get("devices") or {}
    source_basis = freshness_summary.get("source_basis", "persisted_samples")
    for device_id, info in (data.get("app_breakdown") or {}).items():
        freshness = freshness_devices.get(str(device_id)) or freshness_devices.get(device_id) or {}
        for app in info.get("apps") or []:
            rows.append(
                {
                    "Section": "applications",
                    "Device": info.get("device_name"),
                    "Employee": info.get("employee_name"),
                    "Application": app.get("name"),
                    "Category": app.get("category"),
                    "Duration (sec)": app.get("total_seconds"),
                    "Sessions": app.get("sessions"),
                    "Data Basis": source_basis,
                    "Freshness State": freshness.get("freshness_state"),
                    "Last Sample At": freshness.get("last_sample_at"),
                    "Coverage %": freshness.get("coverage_pct"),
                    "Sample Count": freshness.get("sample_count"),
                    "Report Eligible": "Yes" if freshness.get("report_eligible") else "No",
                }
            )
    for category, total_seconds in (data.get("category_totals") or {}).items():
        rows.append({"Section": "category_totals", "Category": category, "Value": total_seconds})
    for device_id, metrics in (data.get("activity_summary") or {}).items():
        device_name = device_id
        employee_name = ""
        metric_items = metrics.items() if isinstance(metrics, dict) else []
        if isinstance(metrics, dict):
            device_name = metrics.get("device_name") or device_id
            employee_name = metrics.get("employee_name") or ""
            metric_items = [
                (metric_name, metric_value)
                for metric_name, metric_value in metrics.items()
                if metric_name not in {"device_name", "employee_name"}
            ]
        for metric_name, metric_value in metric_items:
            rows.append(
                {
                    "Section": "activity_summary",
                    "Device": device_name,
                    "Employee": employee_name,
                    "Metric": metric_name,
                    "Value": metric_value,
                }
            )
    totals = freshness_summary.get("totals") or {}
    for metric_name, metric_value in totals.items():
        rows.append({"Section": "freshness_totals", "Metric": metric_name, "Value": metric_value})
    return headers, rows


def _flatten_network(data):
    headers = [
        "Section",
        "Device",
        "Interface",
        "Timestamp",
        "RX (bps)",
        "TX (bps)",
        "RX Util %",
        "TX Util %",
        "Metric",
        "Value",
    ]
    rows = []
    for _, info in (data.get("bandwidth") or {}).items():
        for point in info.get("points") or []:
            rows.append(
                {
                    "Section": "bandwidth",
                    "Device": info.get("device_name"),
                    "Interface": info.get("interface"),
                    "Timestamp": point.get("ts"),
                    "RX (bps)": point.get("rx_bps"),
                    "TX (bps)": point.get("tx_bps"),
                    "RX Util %": point.get("rx_util"),
                    "TX Util %": point.get("tx_util"),
                }
            )
    for summary in data.get("uptime_summary") or []:
        rows.append(
            {
                "Section": "uptime_summary",
                "Device": summary.get("device_name"),
                "Metric": "avg_uptime",
                "Value": summary.get("avg_uptime"),
            }
        )
        rows.append(
            {
                "Section": "uptime_summary",
                "Device": summary.get("device_name"),
                "Metric": "avg_latency_ms",
                "Value": summary.get("avg_latency_ms"),
            }
        )
        rows.append(
            {
                "Section": "uptime_summary",
                "Device": summary.get("device_name"),
                "Metric": "avg_packet_loss",
                "Value": summary.get("avg_packet_loss"),
            }
        )
    mttr = data.get("mttr") or {}
    rows.append({"Section": "mttr", "Metric": "seconds", "Value": mttr.get("seconds")})
    rows.append({"Section": "mttr", "Metric": "human", "Value": mttr.get("human")})
    rows.append({"Section": "mttr", "Metric": "total_incidents", "Value": mttr.get("total_incidents")})
    return headers, rows


def _flatten_alerts(data):
    headers = [
        "Section",
        "Timestamp",
        "Device",
        "IP",
        "Severity",
        "Type",
        "Message",
        "Acknowledged",
        "Resolved",
        "Label",
        "Value",
    ]
    rows = []
    for alert in data.get("alerts") or []:
        rows.append(
            {
                "Section": "alerts",
                "Timestamp": alert.get("timestamp"),
                "Device": alert.get("device_name"),
                "IP": alert.get("device_ip"),
                "Severity": alert.get("severity"),
                "Type": alert.get("event_type"),
                "Message": alert.get("message"),
                "Acknowledged": "Yes" if alert.get("is_acknowledged") else "No",
                "Resolved": "Yes" if alert.get("resolved") else "No",
            }
        )
    for day, severities in (data.get("daily_trend") or {}).items():
        for severity, count in (severities or {}).items():
            rows.append({"Section": "daily_trend", "Timestamp": day, "Severity": severity, "Value": count})
    for severity, count in (data.get("severity_breakdown") or {}).items():
        rows.append({"Section": "severity_breakdown", "Severity": severity, "Value": count})
    for device in data.get("top_alerted_devices") or []:
        rows.append(
            {
                "Section": "top_alerted_devices",
                "Device": device.get("device_name"),
                "IP": device.get("device_ip"),
                "Value": device.get("alert_count"),
            }
        )
    tta = data.get("tta") or {}
    ttr = data.get("ttr") or {}
    rows.append({"Section": "tta", "Label": "seconds", "Value": tta.get("seconds")})
    rows.append({"Section": "tta", "Label": "human", "Value": tta.get("human")})
    rows.append({"Section": "ttr", "Label": "seconds", "Value": ttr.get("seconds")})
    rows.append({"Section": "ttr", "Label": "human", "Value": ttr.get("human")})
    return headers, rows


def _flatten_maintenance_availability(data):
    headers = ["Section", "Device", "IP", "Start", "End", "Metric", "Value", "Status", "Reason"]
    rows = []
    for window in data.get("scheduled_windows") or []:
        rows.append(
            {
                "Section": "scheduled_windows",
                "Device": window.get("device_name"),
                "IP": window.get("device_ip"),
                "Start": window.get("start_time"),
                "End": window.get("end_time"),
                "Reason": window.get("reason"),
                "Status": "active" if window.get("is_active") else "inactive",
            }
        )
    for device in data.get("maintenance_devices") or []:
        rows.append({"Section": "maintenance_devices", "Device": device.get("device_name"), "IP": device.get("device_ip"), "Status": "maintenance"})
    for item in data.get("downtime_leaders") or []:
        rows.append(
            {
                "Section": "downtime_leaders",
                "Device": item.get("device_name"),
                "IP": item.get("device_ip"),
                "Metric": "availability_pct",
                "Value": item.get("availability_pct"),
            }
        )
    for item in data.get("tracked_instability") or []:
        rows.append(
            {
                "Section": "tracked_instability",
                "Device": item.get("device_name"),
                "Metric": "offline_events",
                "Value": item.get("offline_events"),
            }
        )
        rows.append(
            {
                "Section": "tracked_instability",
                "Device": item.get("device_name"),
                "Metric": "degraded_events",
                "Value": item.get("degraded_events"),
            }
        )
    for metric_name, metric_value in (data.get("summary") or {}).items():
        rows.append({"Section": "summary", "Metric": metric_name, "Value": metric_value})
    return headers, rows


def _flatten_security_compliance(data):
    headers = ["Section", "Timestamp", "Device", "Severity", "Type", "Metric", "Value", "Detail"]
    rows = []
    for metric_name, metric_value in (data.get("summary") or {}).items():
        rows.append({"Section": "summary", "Metric": metric_name, "Value": metric_value})
    for event in data.get("recent_alerts") or []:
        rows.append(
            {
                "Section": "recent_alerts",
                "Timestamp": event.get("timestamp"),
                "Device": event.get("device_name"),
                "Severity": event.get("severity"),
                "Type": event.get("event_type"),
                "Detail": event.get("message"),
            }
        )
    for entry in data.get("recent_audit_log") or []:
        rows.append(
            {
                "Section": "audit_log",
                "Timestamp": entry.get("timestamp"),
                "Type": entry.get("action"),
                "Detail": entry.get("description"),
                "Value": entry.get("entity_id"),
            }
        )
    for item in data.get("restricted_site_violations") or []:
        rows.append(
            {
                "Section": "restricted_site_violations",
                "Timestamp": item.get("observed_at_utc"),
                "Device": item.get("device_name"),
                "Metric": item.get("domain"),
                "Value": item.get("count"),
            }
        )
    for severity, count in (data.get("integrity_breakdown") or {}).items():
        rows.append({"Section": "integrity_breakdown", "Severity": severity, "Value": count})
    for item in data.get("threshold_breaches") or []:
        rows.append(
            {
                "Section": "threshold_breaches",
                "Device": item.get("device_name"),
                "Metric": item.get("metric_key"),
                "Value": item.get("breach_streak"),
                "Detail": item.get("last_state"),
            }
        )
    return headers, rows


def _flatten_inventory_assets(data):
    headers = ["Section", "Name", "Device", "IP", "Metric", "Value", "Type", "Status"]
    rows = []
    for metric_name, metric_value in (data.get("summary") or {}).items():
        rows.append({"Section": "summary", "Metric": metric_name, "Value": metric_value})
    for item in data.get("inventory_devices") or []:
        rows.append(
            {
                "Section": "inventory_devices",
                "Name": item.get("device_name"),
                "IP": item.get("device_ip"),
                "Type": item.get("device_type"),
            }
        )
    for item in data.get("tracked_devices") or []:
        rows.append(
            {
                "Section": "tracked_devices",
                "Name": item.get("device_name"),
                "IP": item.get("ip_address"),
                "Status": item.get("availability_status"),
            }
        )
    for item in data.get("active_links") or []:
        rows.append(
            {
                "Section": "active_links",
                "Device": item.get("device_name"),
                "Name": item.get("tracked_device_name"),
                "Metric": "confidence",
                "Value": item.get("confidence"),
                "Status": item.get("link_source"),
            }
        )
    for item in data.get("pending_candidates") or []:
        rows.append(
            {
                "Section": "pending_candidates",
                "Device": item.get("device_name"),
                "Name": item.get("tracked_device_name"),
                "Metric": "candidate_score",
                "Value": item.get("candidate_score"),
                "Status": item.get("status"),
            }
        )
    return headers, rows


def _flatten_tracking_operations(data):
    headers = [
        "Section",
        "Device",
        "Metric",
        "Value",
        "Application",
        "Category",
        "Timestamp",
        "Status",
        "Last Sample At",
        "Sample Count",
        "Report Eligible",
    ]
    rows = []
    for metric_name, metric_value in (data.get("summary") or {}).items():
        rows.append({"Section": "summary", "Metric": metric_name, "Value": metric_value})
    for item in data.get("device_freshness") or []:
        rows.append(
            {
                "Section": "device_freshness",
                "Device": item.get("device_name"),
                "Metric": "coverage_pct",
                "Value": item.get("coverage_pct"),
                "Status": item.get("freshness_state"),
                "Last Sample At": item.get("last_sample_at"),
                "Sample Count": item.get("sample_count"),
                "Report Eligible": "Yes" if item.get("report_eligible") else "No",
            }
        )
    for item in data.get("top_applications") or []:
        rows.append(
            {
                "Section": "top_applications",
                "Device": item.get("device_name"),
                "Application": item.get("application_name"),
                "Category": item.get("category"),
                "Value": item.get("total_seconds"),
            }
        )
    for item in data.get("activity_totals") or []:
        rows.append(
            {
                "Section": "activity_totals",
                "Device": item.get("device_name"),
                "Metric": item.get("activity_type"),
                "Value": item.get("total_events"),
            }
        )
    for item in data.get("availability_breakdown") or []:
        rows.append(
            {
                "Section": "availability_breakdown",
                "Device": item.get("device_name"),
                "Metric": "offline_events",
                "Value": item.get("offline_events"),
            }
        )
        rows.append(
            {
                "Section": "availability_breakdown",
                "Device": item.get("device_name"),
                "Metric": "degraded_events",
                "Value": item.get("degraded_events"),
            }
        )
        rows.append(
            {
                "Section": "availability_breakdown",
                "Device": item.get("device_name"),
                "Metric": "online_events",
                "Value": item.get("online_events"),
            }
        )
    for severity, count in (data.get("integrity_breakdown") or {}).items():
        rows.append({"Section": "integrity_breakdown", "Metric": severity, "Value": count})
    return headers, rows


def _flatten_printer_operations(data):
    headers = ["Section", "Device", "Metric", "Value", "Timestamp", "Status", "User"]
    rows = []
    for metric_name, metric_value in (data.get("summary") or {}).items():
        rows.append({"Section": "summary", "Metric": metric_name, "Value": metric_value})
    for metric_name, metric_value in (data.get("promotion_triggers") or {}).items():
        rows.append({"Section": "promotion_triggers", "Metric": metric_name, "Value": metric_value})
    for item in data.get("printer_status") or []:
        rows.append(
            {
                "Section": "printer_status",
                "Device": item.get("device_name"),
                "Timestamp": item.get("timestamp"),
                "Status": item.get("status"),
                "Metric": "page_count_total",
                "Value": item.get("page_count_total"),
            }
        )
    for item in data.get("print_volume") or []:
        rows.append(
            {
                "Section": "print_volume",
                "Device": item.get("printer_name"),
                "Metric": "job_count",
                "Value": item.get("job_count"),
                "User": item.get("user_account"),
            }
        )
        rows.append(
            {
                "Section": "print_volume",
                "Device": item.get("printer_name"),
                "Metric": "total_pages",
                "Value": item.get("total_pages"),
                "User": item.get("user_account"),
            }
        )
    return headers, rows


def export_to_csv(report_data, report_type):
    headers, rows = _flatten_report_rows(report_data, report_type)
    display_headers = [_COLUMN_DISPLAY_NAMES.get(h, h) for h in headers]
    text_buf = io.StringIO()
    _write_brand_header_csv(text_buf, report_data, report_type)
    writer = csv.DictWriter(text_buf, fieldnames=display_headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        display_row = {
            _COLUMN_DISPLAY_NAMES.get(h, h): _sanitize_export_value(row.get(h, ""))
            for h in headers
        }
        writer.writerow(display_row)
    # Prepend UTF-8 BOM so Excel opens the file without encoding corruption
    buf = io.BytesIO(b"\xef\xbb\xbf" + text_buf.getvalue().encode("utf-8"))
    buf.seek(0)
    return buf


def export_to_excel(report_data, report_type):
    headers, rows = _flatten_report_rows(report_data, report_type)
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_brand_header_xlsx(ws_summary, report_data, report_type, 2)
    ws_summary.append([""])  # blank separator — row 5
    for label, value in _summary_pairs(report_data, report_type, len(rows)):
        ws_summary.append([label, _excel_cell_value(value)])
    if report_type == "productivity":
        freshness_summary = report_data.get("freshness_summary") or {}
        totals = freshness_summary.get("totals") or {}
        ws_summary.append(["Data Basis", freshness_summary.get("source_basis", "persisted_samples")])
        ws_summary.append(["Fresh Devices", totals.get("fresh_devices", 0)])
        ws_summary.append(["Stale Devices", totals.get("stale_devices", 0)])
        ws_summary.append(["Empty Devices", totals.get("empty_devices", 0)])
    # Start styling from row 6 (rows 1-4 are brand header, row 5 is blank separator)
    for row in ws_summary.iter_rows(min_row=6, max_row=ws_summary.max_row, max_col=2):
        row[0].font = LABEL_FONT
        row[1].font = DATA_FONT
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        _apply_excel_number_format(row[1], row[0].value)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 60

    custom_builders = {
        "executive": _build_executive_excel_sheets,
        "network": _build_network_excel_sheets,
        "alerts": _build_alerts_excel_sheets,
    }
    builder = custom_builders.get(report_type)
    if builder is not None:
        builder(wb, report_data)

    ws_data = wb.create_sheet(title="Data")
    _write_brand_header_xlsx(ws_data, report_data, report_type, len(headers))
    ws_data.append([""])  # blank separator — row 5
    _write_table_sheet(ws_data, headers, rows, start_row=6)

    if builder is None:
        ordered_sections, grouped_rows = _group_rows_by_section(rows)
        for section in ordered_sections:
            sheet_name = _section_sheet_name(section, wb.sheetnames)
            ws_section = wb.create_sheet(title=sheet_name)
            section_rows = grouped_rows.get(section) or []
            _write_brand_header_xlsx(ws_section, report_data, report_type, len(headers))
            ws_section.append([""])  # blank separator — row 5
            _write_table_sheet(ws_section, headers, section_rows, start_row=6)

    warnings = (report_data.get("meta") or {}).get("completeness_warnings") or []
    if warnings:
        ws_warnings = wb.create_sheet(title=_section_sheet_name("Warnings", wb.sheetnames))
        warning_headers = ["Report Type", "Generated At", "Warning"]
        warning_rows = [
            {
                "Report Type": report_type.replace("-", " ").title(),
                "Generated At": (report_data.get("meta") or {}).get("generated_at") or _utc_label(),
                "Warning": warning,
            }
            for warning in warnings
        ]
        _write_titled_table_sheet(ws_warnings, "Completeness Warnings", warning_headers, warning_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _escape_pdf_text(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _export_to_pdf_fallback(report_data, report_type):
    headers, rows = _flatten_report_rows(report_data, report_type)
    preview_rows = rows[:35]
    lines = [
        f"{report_type.replace('-', ' ').title()} Report",
        f"Generated {_utc_label()}",
    ]
    meta = report_data.get("meta") or {}
    if meta:
        lines.append(f"Scope {meta.get('scope_type') or 'global'}:{meta.get('scope_id')}")
        lines.append(f"Freshness {meta.get('freshness_state')}")
    lines.append("")
    lines.append(" | ".join(headers))
    for row in preview_rows:
        lines.append(" | ".join(str(_sanitize_export_value(row.get(header, "")))[:40] for header in headers))

    content_lines = []
    for idx, line in enumerate(lines):
        operator = "Td" if idx == 0 else "T*"
        if idx == 0:
            content_lines.append(f"BT /F1 10 Tf 36 800 {operator} ({_escape_pdf_text(line)}) Tj")
        else:
            content_lines.append(f"({_escape_pdf_text(line)}) Tj {operator}")
    content_lines.append("ET")
    content = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects = []
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj")
    objects.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj")
    objects.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj")
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj")
    objects.append(f"5 0 obj << /Length {len(content)} >> stream\n".encode("ascii") + content + b"\nendstream endobj")

    pdf = io.BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(pdf.tell())
        pdf.write(obj)
        pdf.write(b"\n")
    xref_offset = pdf.tell()
    pdf.write(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.write(
        (
            "trailer << /Size {size} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".format(
                size=len(offsets),
                xref=xref_offset,
            )
        ).encode("ascii")
    )
    pdf.seek(0)
    return pdf


def export_to_pdf(report_data, report_type):
    if not report_data:
        report_data = {}

    try:
        from reportlab.lib import colors  # noqa: F401
    except ImportError:
        return _export_to_pdf_fallback(report_data, report_type)

    builder = _PDF_BUILDERS.get(report_type)
    if builder:
        try:
            return builder(report_data, report_type)
        except Exception:
            logger.exception('[EXPORT] PDF builder failed for %s, using fallback', report_type)
            return _export_to_pdf_fallback(report_data, report_type)

    # Generic path for unmapped report types
    try:
        return _pdf_generic(report_data, report_type)
    except Exception:
        logger.exception('[EXPORT] ReportLab rendering failed for %s, using fallback', report_type)
        return _export_to_pdf_fallback(report_data, report_type)


def _pdf_generic(report_data, report_type):
    """Generic PDF builder for report types without a dedicated builder."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    headers, rows = _flatten_report_rows(report_data, report_type)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"{report_type.replace('-', ' ').title()} Report", styles["Title"]),
        Spacer(1, 8),
    ]
    period = report_data.get("period") or {}
    meta = report_data.get("meta") or {}
    story.append(Paragraph(f"Period: {period.get('start', '')} to {period.get('end', '')}", styles["Normal"]))
    story.append(Paragraph(f"Generated: {_utc_label()}", styles["Normal"]))
    if meta:
        story.append(Paragraph(f"Scope: {meta.get('scope_type', 'global')} / {meta.get('scope_id', '')}", styles["Normal"]))
        story.append(Paragraph(f"Freshness: {meta.get('freshness_state', '')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if not rows:
        story.append(Paragraph("No data available for this period.", styles["Normal"]))
    else:
        table_rows = [headers]
        for row in rows[:250]:
            table_rows.append([str(_sanitize_export_value(row.get(header, ""))) for header in headers])

        table = Table(table_rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2A4A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E0")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf


# ── Per-type PDF builders ────────────────────────────────────────────────────

def _pdf_builder_base(report_data, report_type, title):
    """Shared setup for branded PDF builders. Returns (doc, story, styles, buf)."""
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
    from services.enterprise_pdf_service import (
        base_table_style, section_heading, normal_paragraph, PageFooter,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    gen_at = _utc_label()
    period = report_data.get("period") or {}

    story = [
        section_heading(title, styles),
        normal_paragraph(
            f"Period: {period.get('start', 'N/A')} to {period.get('end', 'N/A')} &nbsp;|&nbsp; Generated: {gen_at}",
            styles,
        ),
        Spacer(1, 10),
    ]
    footer = PageFooter(title, gen_at)
    return doc, story, styles, buf, footer


def _pdf_executive(report_data, report_type):
    from reportlab.platypus import Spacer, Table
    from services.enterprise_pdf_service import base_table_style, normal_paragraph

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Executive Summary Report")

    # KPI summary
    story.append(normal_paragraph(
        f"Fleet Availability: <b>{report_data.get('uptime_score', '—')}%</b> &nbsp;|&nbsp; "
        f"Total Devices: <b>{report_data.get('total_devices', '—')}</b> &nbsp;|&nbsp; "
        f"Avg Latency: <b>{report_data.get('avg_latency', '—')}ms</b>",
        styles,
    ))
    story.append(Spacer(1, 8))

    # Problematic devices table
    headers, rows = _flatten_report_rows(report_data, report_type)
    if not rows:
        story.append(normal_paragraph("No data available for this period.", styles))
    else:
        table_data = [headers]
        for row in rows[:250]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


def _pdf_operational(report_data, report_type):
    from reportlab.platypus import Table
    from services.enterprise_pdf_service import base_table_style, normal_paragraph

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Operational Report")

    # Audit log table
    headers, rows = _flatten_report_rows(report_data, report_type)
    if not rows:
        story.append(normal_paragraph("No data available for this period.", styles))
    else:
        table_data = [headers]
        for row in rows[:250]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


def _pdf_device_health(report_data, report_type):
    from reportlab.platypus import Table
    from services.enterprise_pdf_service import base_table_style, normal_paragraph

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Device Health Report")

    headers, rows = _flatten_report_rows(report_data, report_type)
    if not rows:
        story.append(normal_paragraph("No data available for this period.", styles))
    else:
        table_data = [headers]
        for row in rows[:250]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


def _pdf_cell_value(val, col_name=""):
    """Format a cell value for PDF display. Timestamps → IST, rest → sanitized."""
    if val is None:
        return "—"
    col_lower = col_name.lower()
    # Detect timestamp columns by name or ISO pattern
    if col_lower in ("timestamp", "observed_at_utc", "last_seen", "resolved_at", "acknowledged_at",
                      "created_at", "detected_at", "last_violation", "last_detected"):
        return _fmt_ist(val)
    if isinstance(val, str) and len(val) > 18 and "T" in val and val[4] == "-":
        return _fmt_ist(val)  # Looks like ISO 8601
    return str(_sanitize_export_value(val))


def _pdf_alerts(report_data, report_type):
    from reportlab.platypus import Spacer, Table, Paragraph as RLParagraph
    from reportlab.lib.styles import ParagraphStyle
    from services.enterprise_pdf_service import (
        base_table_style, normal_paragraph, section_heading, hex_color,
        _build_narrative_section,
    )

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Alert History Report")

    # ── Narrative block (Master Spec) ─────────────────────────────────────
    narrative = report_data.get("narrative")
    story.extend(_build_narrative_section(narrative, styles))

    # ── Intelligence annotations ──────────────────────────────────────────
    annotations = report_data.get("intelligence_annotations", [])
    for a in annotations[:5]:
        sev_color = "#ef4444" if a.get("severity") == "critical" else "#f59e0b"
        story.append(RLParagraph(
            f'<font color="{sev_color}">&#x25CF;</font> <b>{a.get("text", "")}</b>'
            f'{" — " + a.get("action", "") if a.get("action") else ""}',
            ParagraphStyle('annotation', parent=styles['Normal'], fontSize=8, leading=11,
                           spaceBefore=2, spaceAfter=2),
        ))

    # ── Severity breakdown ────────────────────────────────────────────────
    breakdown = report_data.get("severity_breakdown") or {}
    if breakdown:
        summary_parts = [f"{k}: {v}" for k, v in breakdown.items()]
        story.append(normal_paragraph(f"Severity Breakdown: {', '.join(summary_parts)}", styles))
        story.append(Spacer(1, 4))

    # ── Alert type breakdown (PR 18) ──────────────────────────────────────
    type_breakdown = report_data.get("alert_type_breakdown", [])
    if type_breakdown:
        story.append(normal_paragraph("<b>Alert Type Breakdown</b>", styles))
        tb_data = [["Type", "Count", "% of Total"]]
        for tb in type_breakdown:
            tb_data.append([tb.get("type", "—"), str(tb.get("count", 0)), f"{tb.get('pct_of_total', 0)}%"])
        t = Table(tb_data, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)
        story.append(Spacer(1, 6))

    # ── Top alerted devices ───────────────────────────────────────────────
    top_devices = report_data.get("top_alerted_devices", [])
    if top_devices:
        story.append(normal_paragraph("<b>Most Impacted Devices</b>", styles))
        td_data = [["Device", "IP", "Alert Count"]]
        for d in top_devices[:10]:
            td_data.append([d.get("device_name", "—"), d.get("device_ip", "—"), str(d.get("alert_count", 0))])
        t = Table(td_data, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)
        story.append(Spacer(1, 6))

    # ── Unresolved aging (PR 18) ──────────────────────────────────────────
    aging = report_data.get("unresolved_aging", {})
    if aging and any(v > 0 for v in aging.values()):
        story.append(normal_paragraph("<b>Unresolved Alert Aging</b>", styles))
        ag_data = [["Age Bucket", "Count"]]
        for bucket, count in aging.items():
            ag_data.append([bucket, str(count)])
        t = Table(ag_data, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)
        story.append(Spacer(1, 6))

    # ── Subnet analysis (PR 18) ───────────────────────────────────────────
    subnets = report_data.get("subnet_analysis", [])
    if subnets:
        story.append(normal_paragraph("<b>Subnet Analysis</b>", styles))
        sn_data = [["Subnet", "Total", "Offline", "Latency", "Pkt Loss", "Devices"]]
        for s in subnets[:10]:
            flag = " *" if s.get("flag") else ""
            sn_data.append([
                s.get("subnet", "—") + flag, str(s.get("total", 0)),
                str(s.get("offline", 0)), str(s.get("latency", 0)),
                str(s.get("pkt_loss", 0)), str(s.get("device_count", 0)),
            ])
        t = Table(sn_data, repeatRows=1)
        t.setStyle(base_table_style())
        story.append(t)
        for s in subnets:
            if s.get("flag"):
                story.append(normal_paragraph(f"* {s['flag']}", styles, color="#EA580C"))
        story.append(Spacer(1, 6))

    # ── Recent alerts table (capped at 20) ────────────────────────────────
    headers, rows = _flatten_report_rows(report_data, report_type)
    total_count = report_data.get("alerts_total_count", len(rows))
    truncated = report_data.get("alerts_truncated", False)

    if not rows:
        story.append(normal_paragraph("No alerts recorded for this period.", styles))
    else:
        story.append(normal_paragraph(
            f"<b>Recent Alerts</b> (showing {len(rows)} of {total_count})"
            + (" — full list available via CSV/XLSX export" if truncated else ""),
            styles,
        ))
        table_data = [headers]
        for row in rows[:50]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


def _pdf_network(report_data, report_type):
    from reportlab.platypus import Spacer, Table
    from services.enterprise_pdf_service import base_table_style, normal_paragraph

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Network Performance Report")

    # MTTR summary
    mttr = report_data.get("mttr") or {}
    if mttr:
        story.append(normal_paragraph(
            f"Mean Time To Resolve: <b>{mttr.get('human', '—')}</b> &nbsp;|&nbsp; "
            f"Resolved Incidents: <b>{mttr.get('total_incidents', 0)}</b>",
            styles,
        ))
        story.append(Spacer(1, 8))

    headers, rows = _flatten_report_rows(report_data, report_type)
    if not rows:
        story.append(normal_paragraph("No data available for this period.", styles))
    else:
        table_data = [headers]
        for row in rows[:250]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


def _pdf_productivity(report_data, report_type):
    from reportlab.platypus import Spacer, Table
    from services.enterprise_pdf_service import base_table_style, normal_paragraph

    doc, story, styles, buf, footer = _pdf_builder_base(report_data, report_type, "Productivity Report")

    # Category summary
    cats = report_data.get("category_totals") or {}
    if cats:
        cat_parts = [f"{k}: {v}" for k, v in cats.items()]
        story.append(normal_paragraph(f"Category Totals: {', '.join(cat_parts)}", styles))
        story.append(Spacer(1, 8))

    headers, rows = _flatten_report_rows(report_data, report_type)
    if not rows:
        story.append(normal_paragraph("No data available for this period.", styles))
    else:
        table_data = [headers]
        for row in rows[:250]:
            table_data.append([_pdf_cell_value(row.get(h, "—"), h) for h in headers])
        table = Table(table_data, repeatRows=1)
        table.setStyle(base_table_style())
        story.append(table)

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf


_PDF_BUILDERS = {
    'executive': _pdf_executive,
    'operational': _pdf_operational,
    'device-health': _pdf_device_health,
    'alerts': _pdf_alerts,
    'network': _pdf_network,
    'productivity': _pdf_productivity,
}


def export_report_buffer(report_data, report_type, export_format=None):
    return export_to_pdf(report_data, report_type)
